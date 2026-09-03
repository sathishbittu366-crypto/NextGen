"""Group 4 — Students API (§7.4).

Source: webapp/routes/students.py — ported to JSON API shape.
Backend for StudentsListPage, StudentFormPage, StudentViewPage which already exist
in frontend/src/pages/students/ and call through frontend/src/api/students.ts.
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, File, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel
from openpyxl import load_workbook

from database import (
    audit, connect, ensure_student_login, mask_aadhaar,
    validate_student, IntegrityError,
)
from field_encryption import decrypt_field, encrypt_field
from sms_app.services.attendance_service import list_semesters
from sms_app.services.student_pdf import build_students_list_pdf
from webapp.photo_upload import PhotoUploadError, save_profile_photo

from api.deps import CurrentUser, get_current_user
from api.envelope import ApiError, ok

router = APIRouter(prefix="/api/students", tags=["students"])

FIELD_SPECS = [
    ("Roll Number", "roll_no"), ("Full Name", "name"), ("Father Name", "father_name"),
    ("Email", "email"), ("Student Phone Number", "phone"), ("Parent Phone Number", "parent_phone"),
    ("Date of Birth (YYYY-MM-DD)", "dob"),
    ("Category", "category"), ("Gender", "gender"), ("Seat Category", "seat_category"),
    ("APAAR ID", "apaar_id"), ("Aadhaar Number", "aadhaar_number"),
    ("Certificates Submitted", "certificates_submitted"), ("Certificates Due", "certificates_due"),
    ("Consultant Name", "consultant_name"), ("Address", "address"),
]

EDUCATION_SPECS = [
    ("10th School Name", "tenth_school"), ("10th Year of Passing", "tenth_year"), ("10th Marks (%)", "tenth_marks"),
    ("12th / Junior College Name", "twelfth_school"), ("12th Year of Passing", "twelfth_year"), ("12th Marks (%)", "twelfth_marks"),
    ("Diploma College Name (if applicable)", "diploma_college"), ("Diploma Year of Passing", "diploma_year"), ("Diploma Marks (%)", "diploma_marks"),
]


def _decrypt_row(row) -> dict:
    """Return a dict with aadhaar_number/apaar_id decrypted.
    Row dictionary is immutable — must copy to dict first.
    """
    if row is None:
        return None
    d = dict(row)
    d["aadhaar_number"] = decrypt_field(d.get("aadhaar_number"))
    d["apaar_id"] = decrypt_field(d.get("apaar_id"))
    return d


def _compute_year_and_batch(d: dict) -> tuple[str, str]:
    roll_no = str(d.get("roll_no") or "").strip().upper()
    sem_id = d.get("current_semester_id")

    # Determine batch from roll number (e.g. 24BT1A6722 -> 2024-2028 Batch)
    batch = ""
    joining_year = None
    if len(roll_no) >= 2 and roll_no[:2].isdigit():
        yy = int(roll_no[:2])
        if 18 <= yy <= 35:
            joining_year = 2000 + yy
            batch = f"{joining_year}-{joining_year + 4} Batch"

    # Determine Year of study
    year = ""
    if sem_id in (1, 2):
        year = "1st Year"
        if not batch: batch = "2026-2030 Batch"
    elif sem_id in (3, 4):
        year = "2nd Year"
        if not batch: batch = "2025-2029 Batch"
    elif sem_id in (5, 6):
        year = "3rd Year"
        if not batch: batch = "2024-2028 Batch"
    elif sem_id in (7, 8):
        year = "4th Year"
        if not batch: batch = "2023-2027 Batch"
    elif joining_year:
        diff = 2026 - joining_year + 1
        if diff <= 1: year = "1st Year"
        elif diff == 2: year = "2nd Year"
        elif diff == 3: year = "3rd Year"
        else: year = "4th Year"
    else:
        year = "1st Year"
        batch = "2026-2030 Batch"

    return year, batch


def _serialize_list_row(row) -> dict:
    """For the list endpoint — decrypt then mask aadhaar."""
    d = _decrypt_row(row)
    year, batch = _compute_year_and_batch(d)
    return {
        "id": d["id"],
        "roll_no": d["roll_no"],
        "name": d["name"],
        "email": d.get("email") or "",
        "phone": d.get("phone") or "",
        "parent_phone": d.get("parent_phone") or "",
        "gender": d.get("gender") or "",
        "category": d.get("category") or "",
        "seat_category": d.get("seat_category") or "",
        "current_semester_id": d.get("current_semester_id"),
        "aadhaar_masked": mask_aadhaar(d.get("aadhaar_number")),
        "year_of_study": year,
        "batch": batch,
        "active": bool(d["active"]),
        "photo_path": d.get("photo_path"),
    }


def _serialize_full(row) -> dict:
    """Full student record — aadhaar_number/apaar_id returned decrypted."""
    d = _decrypt_row(row)
    year, batch = _compute_year_and_batch(d)
    res = {k: (d.get(k) or "") if isinstance(d.get(k), str) or d.get(k) is None else d[k]
           for k in d if k not in ("password",)}
    res["year_of_study"] = year
    res["batch"] = batch
    return res


def _get_user_hod_username(username: str) -> str | None:
    with connect() as c:
        row = c.execute("SELECT role, hod_username, department FROM users WHERE username=%s", (username,)).fetchone()
        if not row:
            return None
        if row["role"] == "HOD":
            return username
        if row.get("hod_username"):
            return row["hod_username"]
        from database import resolve_hod_for_department
        dept_hod = resolve_hod_for_department(c, row.get("department") or "CSD")
        return dept_hod or row.get("hod_username") or username


def _resolve_hod_for_student(user: CurrentUser, requested: str | None = None) -> str:
    if user.role == "HOD":
        return user.username
    candidate = (requested or "").strip()
    with connect() as c:
        if candidate:
            row = c.execute("SELECT username FROM users WHERE username=%s AND role='HOD' AND active=1", (candidate,)).fetchone()
            if not row:
                raise ApiError("Selected HOD is not an active HOD account", 400, "VALIDATION_ERROR")
            return candidate
        from database import resolve_hod_for_department
        dept_hod = resolve_hod_for_department(c, "CSD")
        if dept_hod:
            return dept_hod
    return user.username


def _student_scope_sql(user: CurrentUser) -> tuple[str, list[Any]]:
    if user.role == "ADMIN" or user.username == "admin":
        return "", []
    hod = user.username if user.role == "HOD" else _get_user_hod_username(user.username)
    if hod:
        return " AND (hod_username=? OR hod_username IS NULL)", [hod]
    return " AND 1=0", []


@router.get("/semesters")
async def student_semesters(user: CurrentUser = Depends(get_current_user)):
    if user.role == "STUDENT":
        raise ApiError("Access denied", 403, "FORBIDDEN")
    semesters = [dict(s) for s in list_semesters()]
    return ok(semesters)


@router.get("")
async def students_list(
    q: str = "",
    status: str = "Active",
    semester_id: int | None = None,
    user: CurrentUser = Depends(get_current_user),
):
    if user.role == "STUDENT":
        raise ApiError("Access denied", 403, "FORBIDDEN")
    like = f"%{q.strip()}%"
    sql = (
        "SELECT * FROM students WHERE department='CSD' "
        "AND (name LIKE ? OR roll_no LIKE ? OR email LIKE ? OR phone LIKE ? OR parent_phone LIKE ?)"
    )
    args: list[Any] = [like, like, like, like, like]
    scope_sql, scope_args = _student_scope_sql(user)
    sql += scope_sql
    args.extend(scope_args)
    if status != "All":
        sql += " AND active=?"
        args.append(1 if status == "Active" else 0)
    if semester_id:
        sql += " AND current_semester_id=?"
        args.append(semester_id)
    sql += " ORDER BY UPPER(roll_no) ASC, name ASC"
    with connect() as c:
        rows = c.execute(sql, args).fetchall()
    return ok([_serialize_list_row(r) for r in rows])


@router.get("/pdf")
async def students_pdf(
    q: str = "",
    status: str = "Active",
    semester_id: int | None = None,
    year: str | None = None,
    user: CurrentUser = Depends(get_current_user),
):
    """Generates official ReportLab PDF for the student list / nominal roll.
    Strictly restricted to HOD and ADMIN roles.
    """
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")

    like = f"%{q.strip()}%"
    sql = (
        "SELECT * FROM students WHERE department='CSD' "
        "AND (name LIKE ? OR roll_no LIKE ? OR email LIKE ? OR phone LIKE ? OR parent_phone LIKE ?)"
    )
    args: list[Any] = [like, like, like, like, like]
    scope_sql, scope_args = _student_scope_sql(user)
    sql += scope_sql
    args.extend(scope_args)

    if status != "All":
        sql += " AND active=?"
        args.append(1 if status == "Active" else 0)
    if semester_id:
        sql += " AND current_semester_id=?"
        args.append(semester_id)

    sql += " ORDER BY UPPER(roll_no) ASC, name ASC"

    with connect() as c:
        rows = c.execute(sql, args).fetchall()
        sem_row = None
        if semester_id:
            sem_row = c.execute("SELECT code, name FROM academic_semesters WHERE id=?", (semester_id,)).fetchone()

    serialized = [_serialize_list_row(r) for r in rows]

    YEAR_LABELS = {"1": "1st Year", "2": "2nd Year", "3": "3rd Year", "4": "4th Year"}
    if year and year in YEAR_LABELS:
        serialized = [r for r in serialized if r.get("year_of_study") == YEAR_LABELS[year]]

    if sem_row:
        semester_name = f"{sem_row['name']} ({sem_row['code']})"
    elif semester_id:
        semester_name = f"Semester {semester_id}"
    elif year and year in YEAR_LABELS:
        semester_name = f"{YEAR_LABELS[year]} (All Semesters)"
    else:
        semester_name = "All Semesters"

    batches = {r.get("batch") for r in serialized if r.get("batch")}
    if len(batches) == 1:
        batch_name = list(batches)[0]
    elif len(batches) > 1:
        batch_name = ", ".join(sorted(batches))
    else:
        batch_name = "All Batches"

    now_str = datetime.now().strftime("%d-%m-%Y %I:%M %p")
    meta = {
        "semester_name": semester_name,
        "batch": batch_name,
        "total_students": len(serialized),
        "generated_on": now_str,
        "generated_by": "NextGen SMS",
        "department": "CSE (DATA SCIENCE)",
    }

    pdf_bytes = build_students_list_pdf(serialized, meta)

    clean_sem = (sem_row["code"] if sem_row else (f"Year_{year}" if year else "All_Semesters")).replace(" ", "_").replace("/", "-")
    filename = f"Student_List_{clean_sem}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{filename}"'},
    )


@router.get("/new")
async def student_new(user: CurrentUser = Depends(get_current_user)):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD access only", 403, "FORBIDDEN")
    semesters = [dict(s) for s in list_semesters()]
    return ok({"student": None, "semesters": semesters})


@router.get("/{student_id}/edit")
async def student_edit_data(student_id: int, user: CurrentUser = Depends(get_current_user)):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD access only", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT * FROM students WHERE id=? AND department='CSD'", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT * FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            hod = _get_user_hod_username(user.username)
            row = c.execute("SELECT * FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, hod)).fetchone() if hod else None
    if not row:
        raise ApiError("Student not found", 404, "NOT_FOUND")
    # CORRUPTION TRAP: must decrypt before returning — see §4.4 and webapp/routes/students.py
    student = _serialize_full(row)
    semesters = [dict(s) for s in list_semesters()]
    return ok({"student": student, "semesters": semesters})


@router.get("/{student_id}")
async def student_view(student_id: int, user: CurrentUser = Depends(get_current_user)):
    if user.role == "STUDENT":
        raise ApiError("Access denied", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT * FROM students WHERE id=? AND department='CSD'", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT * FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            hod = _get_user_hod_username(user.username)
            row = c.execute("SELECT * FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, hod)).fetchone() if hod else None
        semester = None
        if row and row["current_semester_id"]:
            sem_row = c.execute(
                "SELECT code, name FROM academic_semesters WHERE id=?",
                (row["current_semester_id"],)
            ).fetchone()
            if sem_row:
                semester = dict(sem_row)
    if not row:
        raise ApiError("Student not found", 404, "NOT_FOUND")
    # HOD detail view — aadhaar returned decrypted (§4.4 exception)
    student = _serialize_full(row)

    # Subject-wise attendance calculation
    from sms_app.services.attendance_service import student_subject_attendance, attendance_pct_band
    att_rows = student_subject_attendance(row["roll_no"])
    subjects_attendance = []
    total_classes = 0
    total_present = 0
    for ar in att_rows:
        pct, band = attendance_pct_band(ar["present_sessions"], ar["total_sessions"])
        pres = int(ar["present_sessions"] or 0)
        tot = int(ar["total_sessions"] or 0)
        total_classes += tot
        total_present += pres
        subjects_attendance.append({
            "subject_id": ar["subject_id"],
            "subject_code": ar["subject_code"],
            "subject_name": ar["subject_name"],
            "present_sessions": pres,
            "total_sessions": tot,
            "absent_sessions": tot - pres,
            "pct": pct,
            "band": band,
        })
    overall_pct, overall_band = attendance_pct_band(total_present, total_classes)
    attendance_summary = {
        "subjects": subjects_attendance,
        "total_classes": total_classes,
        "total_present": total_present,
        "total_absent": total_classes - total_present,
        "overall_pct": overall_pct,
        "overall_band": overall_band,
    }

    return ok({"student": student, "semester": semester, "attendance": attendance_summary})


@router.get("/{student_id}/attendance")
async def student_attendance(student_id: int, user: CurrentUser = Depends(get_current_user)):
    if user.role == "STUDENT":
        raise ApiError("Access denied", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT roll_no, name FROM students WHERE id=? AND department='CSD'", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT roll_no, name FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            hod = _get_user_hod_username(user.username)
            row = c.execute("SELECT roll_no, name FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, hod)).fetchone() if hod else None
    if not row:
        raise ApiError("Student not found", 404, "NOT_FOUND")

    from sms_app.services.attendance_service import student_subject_attendance, attendance_pct_band, student_subject_session_history
    att_rows = student_subject_attendance(row["roll_no"])
    subjects_attendance = []
    total_classes = 0
    total_present = 0
    for ar in att_rows:
        pct, band = attendance_pct_band(ar["present_sessions"], ar["total_sessions"])
        pres = int(ar["present_sessions"] or 0)
        tot = int(ar["total_sessions"] or 0)
        total_classes += tot
        total_present += pres
        history = student_subject_session_history(row["roll_no"], ar["subject_id"])
        subjects_attendance.append({
            "subject_id": ar["subject_id"],
            "subject_code": ar["subject_code"],
            "subject_name": ar["subject_name"],
            "present_sessions": pres,
            "total_sessions": tot,
            "absent_sessions": tot - pres,
            "pct": pct,
            "band": band,
            "sessions": [
                {
                    "attendance_date": h["attendance_date"],
                    "session_type": h["session_type"],
                    "duration_hours": h["duration_hours"],
                    "status": h["status"],
                }
                for h in history
            ],
        })
    overall_pct, overall_band = attendance_pct_band(total_present, total_classes)
    return ok({
        "roll_no": row["roll_no"],
        "name": row["name"],
        "total_classes": total_classes,
        "total_present": total_present,
        "total_absent": total_classes - total_present,
        "overall_pct": overall_pct,
        "overall_band": overall_band,
        "subjects": subjects_attendance,
    })


def _clean_str(val: str | None) -> str:
    return (val or "").strip()


def _cell_to_str(val: Any) -> str:
    """openpyxl hands back int/float for numeric-looking cells (very
    common for Aadhaar/phone columns unless the column was explicitly
    Text-formatted in Excel) — str(123456789012.0) is '123456789012.0',
    which breaks validate_student's exact-digit-count checks. Strip a
    trailing '.0' from whole-number floats; everything else -> plain str."""
    if val is None:
        return ""
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val).strip()


class StudentBody(BaseModel):
    roll_no: str | None = ""
    name: str | None = ""
    father_name: str | None = ""
    email: str | None = ""
    phone: str | None = ""
    parent_phone: str | None = ""
    dob: str | None = ""
    category: str | None = ""
    gender: str | None = ""
    seat_category: str | None = ""
    apaar_id: str | None = ""
    aadhaar_number: str | None = ""
    certificates_submitted: str | None = ""
    certificates_due: str | None = ""
    consultant_name: str | None = ""
    address: str | None = ""
    tenth_school: str | None = ""
    tenth_year: str | None = ""
    tenth_marks: str | None = ""
    twelfth_school: str | None = ""
    twelfth_year: str | None = ""
    twelfth_marks: str | None = ""
    diploma_college: str | None = ""
    diploma_year: str | None = ""
    diploma_marks: str | None = ""
    current_semester_id: int | None = None
    hod_username: str | None = None


def _body_to_data(body: StudentBody) -> dict:
    return {
        "roll_no": _clean_str(body.roll_no),
        "name": _clean_str(body.name),
        "department": "CSD",
        "email": _clean_str(body.email),
        "phone": _clean_str(body.phone),
        "parent_phone": _clean_str(body.parent_phone),
        "dob": _clean_str(body.dob),
        "address": _clean_str(body.address),
        "father_name": _clean_str(body.father_name),
        "category": _clean_str(body.category),
        "gender": _clean_str(body.gender),
        "seat_category": _clean_str(body.seat_category),
        "apaar_id": _clean_str(body.apaar_id),
        "aadhaar_number": _clean_str(body.aadhaar_number),
        "certificates_submitted": _clean_str(body.certificates_submitted),
        "certificates_due": _clean_str(body.certificates_due),
        "consultant_name": _clean_str(body.consultant_name),
        "tenth_school": _clean_str(body.tenth_school),
        "tenth_year": _clean_str(body.tenth_year),
        "tenth_marks": _clean_str(body.tenth_marks),
        "twelfth_school": _clean_str(body.twelfth_school),
        "twelfth_year": _clean_str(body.twelfth_year),
        "twelfth_marks": _clean_str(body.twelfth_marks),
        "diploma_college": _clean_str(body.diploma_college),
        "diploma_year": _clean_str(body.diploma_year),
        "diploma_marks": _clean_str(body.diploma_marks),
    }


STUDENT_DB_KEYS = [
    "roll_no", "name", "department", "email", "phone", "parent_phone", "dob", "address",
    "father_name", "category", "gender", "seat_category", "apaar_id", "aadhaar_number",
    "certificates_submitted", "certificates_due", "consultant_name",
    "tenth_school", "tenth_year", "tenth_marks",
    "twelfth_school", "twelfth_year", "twelfth_marks",
    "diploma_college", "diploma_year", "diploma_marks",
]


@router.post("")
async def student_create(body: StudentBody, user: CurrentUser = Depends(get_current_user)):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")
    data = _body_to_data(body)
    assigned_hod = _resolve_hod_for_student(user, body.hod_username)
    try:
        validate_student(data)
        if data["dob"]:
            datetime.strptime(data["dob"], "%Y-%m-%d")
        # Encrypt AFTER validation — §4.4 ordering requirement
        data["aadhaar_number"] = encrypt_field(data["aadhaar_number"])
        data["apaar_id"] = encrypt_field(data["apaar_id"])
        with connect() as c:
            c.execute(
                """INSERT INTO students(roll_no,name,department,email,phone,parent_phone,dob,address,father_name,
                   category,gender,seat_category,apaar_id,aadhaar_number,
                   certificates_submitted,certificates_due,consultant_name,
                   tenth_school,tenth_year,tenth_marks,twelfth_school,twelfth_year,twelfth_marks,
                   diploma_college,diploma_year,diploma_marks,current_semester_id,hod_username)
                   VALUES(?,?,?,NULLIF(?,''),?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (*[data[k] for k in STUDENT_DB_KEYS], body.current_semester_id, assigned_hod),
            )
            audit(c, user.username, "CREATE", "student", data["roll_no"])
            # Seed checklist — §7.4
            for item, st in [("Personal details", "Complete"), ("Documents", "Pending"),
                              ("ID card", "Pending"), ("Fees", "Pending"),
                              ("Attendance records", "Available"), ("Marks records", "Available")]:
                c.execute("INSERT IGNORE INTO checklist(roll_no,item,status) VALUES(?,?,?)",
                          (data["roll_no"], item, st))
            new_id = c.execute("SELECT id FROM students WHERE roll_no=?", (data["roll_no"],)).fetchone()["id"]
        username, password = ensure_student_login(data["roll_no"], user.username)
        return ok({"id": new_id, "created_credentials": {"username": username, "password": password}})
    except ValueError as e:
        raise ApiError(str(e), 400, "VALIDATION_ERROR")
    except IntegrityError as e:
        raise ApiError("A student with that roll number or email already exists", 400, "VALIDATION_ERROR")


# — Bulk Import (Excel) —
# Column headers this endpoint understands, matched case-insensitively
# and with surrounding whitespace stripped — mirrors the real sheet
# handed over for 2nd-year seeding. Extra/unknown columns are ignored;
# "Mother's Name" / "Mother's Phone Number" / "Profile photo of student"
# have no matching DB field and are intentionally dropped (per decision:
# no schema change for mother fields, photos skipped, Father's Phone ->
# parent_phone since there's only one generic parent-phone column).
BULK_IMPORT_COLUMN_MAP = {
    "full name of the student": "name",
    "full name of the student (as per ssc)": "name",
    "hallticket": "roll_no",
    "phone no": "phone",
    "student email id": "email",
    "address of the student": "address",
    "aadhaar number": "aadhaar_number",
    "father's name": "father_name",
    "father's phone number": "parent_phone",
}

# All rows from this sheet are seeded into II-I (2nd Year, 1st Semester) —
# matches the "which semester" decision for this specific import batch.
# If a future sheet mixes years, this becomes a per-row lookup instead.
BULK_IMPORT_SEMESTER_CODE = "II-I"


def _normalize_header(h: Any) -> str:
    # collapse internal whitespace too (not just strip ends) — the real
    # sheet's "FULL NAME OF THE STUDENT       (as per ssc)" header has
    # irregular internal spacing that varies by export/copy-paste, so a
    # plain .strip() would silently fail to match on a slightly
    # different run of spaces.
    return " ".join(str(h or "").split()).lower()


def _row_to_import_data(row_map: dict[str, Any]) -> dict:
    """Same shape _body_to_data() produces, built from a spreadsheet row
    instead of a StudentBody — every FIELD_SPECS key present so validate_student()
    and the INSERT below (STUDENT_DB_KEYS) work unmodified. Uses _cell_to_str
    (not _clean_str) since these are raw openpyxl cell values, which may be
    int/float for numeric-looking columns like Aadhaar or phone."""
    return {
        "roll_no": _cell_to_str(row_map.get("roll_no")),
        "name": _cell_to_str(row_map.get("name")),
        "department": "CSD",
        "email": _cell_to_str(row_map.get("email")),
        "phone": _cell_to_str(row_map.get("phone")),
        "parent_phone": _cell_to_str(row_map.get("parent_phone")),
        "dob": "",
        "address": _cell_to_str(row_map.get("address")),
        "father_name": _cell_to_str(row_map.get("father_name")),
        "category": "",
        "gender": "",
        "seat_category": "",
        "apaar_id": "",
        "aadhaar_number": _cell_to_str(row_map.get("aadhaar_number")),
        "certificates_submitted": "",
        "certificates_due": "",
        "consultant_name": "",
        "tenth_school": "",
        "tenth_year": "",
        "tenth_marks": "",
        "twelfth_school": "",
        "twelfth_year": "",
        "twelfth_marks": "",
        "diploma_college": "",
        "diploma_year": "",
        "diploma_marks": "",
    }


@router.post("/bulk-import")
async def student_bulk_import(
    file: UploadFile = File(...),
    user: CurrentUser = Depends(get_current_user),
):
    """Upload an .xlsx sheet -> create a student + login for every row.
    One DB transaction PER ROW (matches student_create's pattern) so a
    single bad row doesn't roll back rows already committed before it —
    the response reports success/skip/error per row instead of an
    all-or-nothing batch result.
    """
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")

    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xlsm")):
        raise ApiError("Upload an .xlsx file exported from Excel", 400, "VALIDATION_ERROR")

    raw = await file.read()
    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise ApiError("Could not read that file — is it a valid .xlsx?", 400, "VALIDATION_ERROR")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise ApiError("The sheet is empty", 400, "VALIDATION_ERROR")

    # Map each spreadsheet column index -> our internal field name, for
    # every header we recognise (see BULK_IMPORT_COLUMN_MAP above).
    col_index_to_field: dict[int, str] = {}
    for idx, header in enumerate(header_row):
        field = BULK_IMPORT_COLUMN_MAP.get(_normalize_header(header))
        if field:
            col_index_to_field[idx] = field

    if "roll_no" not in col_index_to_field.values() or "name" not in col_index_to_field.values():
        raise ApiError(
            "Couldn't find the 'HallTicket' and/or 'FULL NAME OF THE STUDENT' columns — check the sheet's header row",
            400, "VALIDATION_ERROR",
        )

    assigned_hod = _resolve_hod_for_student(user, None)

    with connect() as c:
        sem_row = c.execute(
            "SELECT id FROM academic_semesters WHERE code=%s", (BULK_IMPORT_SEMESTER_CODE,)
        ).fetchone()
    semester_id = sem_row["id"] if sem_row else None

    created: list[dict] = []
    skipped: list[dict] = []
    failed: list[dict] = []

    for row_num, row in enumerate(rows_iter, start=2):  # start=2: row 1 is the header
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue  # blank row — silently skip, not an error

        row_map: dict[str, Any] = {}
        for idx, field in col_index_to_field.items():
            if idx < len(row):
                row_map[field] = row[idx]

        data = _row_to_import_data(row_map)
        roll_no_display = data["roll_no"] or f"(row {row_num})"

        if not data["roll_no"] or not data["name"]:
            failed.append({"row": row_num, "roll_no": roll_no_display, "reason": "Missing roll number or name"})
            continue

        try:
            validate_student(data)
        except ValueError as e:
            failed.append({"row": row_num, "roll_no": roll_no_display, "reason": str(e)})
            continue

        try:
            with connect() as c:
                existing = c.execute(
                    "SELECT id FROM students WHERE roll_no=%s", (data["roll_no"],)
                ).fetchone()
                if existing:
                    skipped.append({"row": row_num, "roll_no": data["roll_no"], "reason": "Roll number already exists"})
                    continue

                enc_data = dict(data)
                enc_data["aadhaar_number"] = encrypt_field(data["aadhaar_number"])
                enc_data["apaar_id"] = encrypt_field(data["apaar_id"])

                c.execute(
                    """INSERT INTO students(roll_no,name,department,email,phone,parent_phone,dob,address,father_name,
                       category,gender,seat_category,apaar_id,aadhaar_number,
                       certificates_submitted,certificates_due,consultant_name,
                       tenth_school,tenth_year,tenth_marks,twelfth_school,twelfth_year,twelfth_marks,
                       diploma_college,diploma_year,diploma_marks,current_semester_id,hod_username)
                       VALUES(?,?,?,NULLIF(?,''),?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (*[enc_data[k] for k in STUDENT_DB_KEYS], semester_id, assigned_hod),
                )
                audit(c, user.username, "CREATE", "student", data["roll_no"])
                for item, st in [("Personal details", "Complete"), ("Documents", "Pending"),
                                  ("ID card", "Pending"), ("Fees", "Pending"),
                                  ("Attendance records", "Available"), ("Marks records", "Available")]:
                    c.execute("INSERT IGNORE INTO checklist(roll_no,item,status) VALUES(?,?,?)",
                              (data["roll_no"], item, st))

            username, password = ensure_student_login(data["roll_no"], user.username)
            created.append({
                "row": row_num,
                "roll_no": data["roll_no"],
                "name": data["name"],
                "username": username,
                "password": password,
            })
        except IntegrityError:
            skipped.append({"row": row_num, "roll_no": data["roll_no"], "reason": "Duplicate roll number or email"})
        except ValueError as e:
            failed.append({"row": row_num, "roll_no": data["roll_no"], "reason": str(e)})
        except Exception as e:
            failed.append({"row": row_num, "roll_no": data["roll_no"], "reason": "Unexpected error — see server logs"})

    return ok({
        "total_rows": len(created) + len(skipped) + len(failed),
        "created_count": len(created),
        "skipped_count": len(skipped),
        "failed_count": len(failed),
        "created": created,
        "skipped": skipped,
        "failed": failed,
    })


@router.patch("/{student_id}")
async def student_update(student_id: int, body: StudentBody, user: CurrentUser = Depends(get_current_user)):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            existing = c.execute("SELECT id,hod_username FROM students WHERE id=? AND department='CSD'", (student_id,)).fetchone()
        elif user.role == "HOD":
            existing = c.execute("SELECT id,hod_username FROM students WHERE id=? AND department='CSD' AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            existing = c.execute("SELECT id,hod_username FROM students WHERE id=? AND department='CSD'", (student_id,)).fetchone()
    if not existing:
        raise ApiError("Student not found", 404, "NOT_FOUND")
    assigned_hod = existing.get("hod_username") if user.role == "HOD" else _resolve_hod_for_student(user, body.hod_username or existing.get("hod_username"))
    data = _body_to_data(body)
    try:
        validate_student(data)
        if data["dob"]:
            datetime.strptime(data["dob"], "%Y-%m-%d")
        # Encrypt AFTER validation — §4.4 ordering requirement
        data["aadhaar_number"] = encrypt_field(data["aadhaar_number"])
        data["apaar_id"] = encrypt_field(data["apaar_id"])
        with connect() as c:
            c.execute(
                """UPDATE students SET roll_no=?,name=?,department=?,email=NULLIF(?,''),phone=?,parent_phone=?,dob=?,
                   address=?,father_name=?,category=?,gender=?,seat_category=?,apaar_id=?,aadhaar_number=?,
                   certificates_submitted=?,certificates_due=?,consultant_name=?,
                   tenth_school=?,tenth_year=?,tenth_marks=?,twelfth_school=?,twelfth_year=?,twelfth_marks=?,
                   diploma_college=?,diploma_year=?,diploma_marks=?,
                   current_semester_id=?,hod_username=?,updated_at=CURRENT_TIMESTAMP WHERE id=?""",
                (*[data[k] for k in STUDENT_DB_KEYS], body.current_semester_id, assigned_hod, student_id),
            )
            audit(c, user.username, "UPDATE", "student", data["roll_no"])
        return ok({"id": student_id, "created_credentials": None})
    except ValueError as e:
        raise ApiError(str(e), 400, "VALIDATION_ERROR")
    except IntegrityError:
        raise ApiError("A student with that roll number or email already exists", 400, "VALIDATION_ERROR")


@router.post("/{student_id}/toggle-status")
async def toggle_status(student_id: int, user: CurrentUser = Depends(get_current_user)):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT * FROM students WHERE id=? AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        if not row:
            raise ApiError("Student not found", 404, "NOT_FOUND")
        new_active = 0 if row["active"] else 1
        c.execute("UPDATE students SET active=?,updated_at=CURRENT_TIMESTAMP WHERE id=?", (new_active, student_id))
        audit(c, user.username, "STATUS", "student", f"{row['roll_no']} -> {new_active}")
    return ok({"active": bool(new_active)})


@router.post("/{student_id}/photo")
async def student_photo(
    student_id: int,
    photo: UploadFile = File(...),
    user: CurrentUser = Depends(get_current_user),
):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT * FROM students WHERE id=? AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
    if not row:
        raise ApiError("Student not found", 404, "NOT_FOUND")
    try:
        path = await save_profile_photo(photo, subdir="students", stem=row["roll_no"])
        with connect() as c:
            c.execute("UPDATE students SET photo_path=?,updated_at=CURRENT_TIMESTAMP WHERE id=?", (path, student_id))
            audit(c, user.username, "PHOTO", "student", row["roll_no"])
        return ok({"photo_path": path})
    except PhotoUploadError as e:
        raise ApiError(str(e), 400, "UPLOAD_ERROR")


@router.post("/{student_id}/photo/delete")
async def student_photo_delete(
    student_id: int,
    user: CurrentUser = Depends(get_current_user),
):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT * FROM students WHERE id=? AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        if not row:
            raise ApiError("Student not found", 404, "NOT_FOUND")
        c.execute("UPDATE students SET photo_path=NULL,updated_at=CURRENT_TIMESTAMP WHERE id=?", (student_id,))
        audit(c, user.username, "PHOTO_DELETE", "student", row["roll_no"])
    return ok({"photo_path": None})


@router.delete("/{student_id}")
async def student_delete(student_id: int, user: CurrentUser = Depends(get_current_user)):
    if user.role not in ("HOD", "ADMIN"):
        raise ApiError("HOD or Admin access only", 403, "FORBIDDEN")
    with connect() as c:
        if user.role == "ADMIN" or user.username == "admin":
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        elif user.role == "HOD":
            row = c.execute("SELECT * FROM students WHERE id=? AND (hod_username=? OR hod_username IS NULL)", (student_id, user.username)).fetchone()
        else:
            row = c.execute("SELECT * FROM students WHERE id=?", (student_id,)).fetchone()
        if not row:
            raise ApiError("Student not found", 404, "NOT_FOUND")
        roll_no = row["roll_no"]
        c.execute("DELETE FROM attendance_records WHERE roll_no=?", (roll_no,))
        c.execute("DELETE FROM sms_queue WHERE roll_no=?", (roll_no,))
        c.execute("DELETE FROM checklist WHERE roll_no=?", (roll_no,))
        c.execute("DELETE FROM users WHERE student_roll_no=? OR username=?", (roll_no, roll_no))
        c.execute("DELETE FROM students WHERE id=?", (student_id,))
        audit(c, user.username, "DELETE", "student", roll_no)
    return ok({"deleted": True, "id": student_id})


# Halted certificate upload (§4.5) — accepts the request, returns success without writing anything
@router.post("/{student_id}/certificate/{doc_type}")
async def certificate_upload(student_id: int, doc_type: str, user: CurrentUser = Depends(get_current_user)):
    """HALTED per §4.5 — certificate upload disabled. Route kept for backward-compat."""
    return ok({"message": "Certificate upload is currently disabled"})